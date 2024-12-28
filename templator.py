import datetime
import importlib
import importlib.util
import os
import subprocess
from typing import Any, Dict

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, \
    Border, \
    Side, \
    Alignment, \
    Protection, \
    Font 
from openpyxl.utils import range_boundaries

from PyPDF2 import PageObject, PdfReader, PdfWriter

from logger import Logger

class Templator:
    @staticmethod
    def start_gen(config: Dict[str, Any]) -> str:
        Logger.print("Start generation process")
        Logger.print(config)
        
        temp_xlsx = Templator.create_xlsx_from_template(
            config.get("report_path"),
            config.get("params" if \
                       config.get("data_was_prepared") else \
                        Templator.dynamic_call(
                            config.get("path_to_py_module"),
                            config.get("params")
                        ),
                        config.get("output"))
        )
        Logger.print(f"Temporary xlsx {temp_xlsx} was generated")
        
        Logger.print("\t Convert xlsx to pdf with libreoffice")
        convert_pdf = Templator.excell_to_pdf(
            temp_xlsx,
            config.get("output"),
            config.get("libreoffice_path")
        )
        Logger.print(f"\t Pdf file {convert_pdf} was created")
        os.remove(temp_xlsx)
        Logger.print(f"\t{temp_xlsx} was deleted")

        dir, old_name = os.path.split(convert_pdf)
        new_name = f"{old_name}_{datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.pdf"
        new_path = os.path.join(dir, new_name)
        Templator.add_metadata_to_pdf(
            input_pdf=convert_pdf,
            output_pdf=new_path,
            watermark=config.get("path_to_watermark"),
            metadata=config.get("metadata")
        )
        os.remove(convert_pdf)
        Logger.print(f"\t{old_name} was renamed as {new_name}")
        return new_path
    
    @staticmethod
    def dynamic_call(
        script_path: str,
        params: Dict[str, Any]
    ) -> Dict[str, Any]:
        Logger.print(f"Calls extended module: {script_path}")
        
        spec = importlib.spec_from_file_location("dynamic_module", script_path)
        if not spec:
            Logger.print(f"cannot load module from {script_path}", level='critical')
            raise ImportError(f"cannot load module from {script_path}")

        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        if not hasattr(module, 'main'):
            Logger.print(f"The script at {script_path} does not have 'main' func", level='critical')
            raise ImportError(f"The script at {script_path} does not have 'main' func")

        result = module.main(params)
        if not isinstance(result, Dict[str, any]):
            Logger.print("The 'main' function must return Dict[str, str] or Dict[str, Dict[str, str[]]]",
                         level='critical')
            raise ValueError("The 'main' function must return Dict[str, str] or Dict[str, Dict[str, str[]]]")
        
        Logger.print(f"Extended module finished with result {result}")
        return result
    
    @staticmethod
    def create_xlsx_from_template(
        template_path:str,
        params:Dict[Any, str],
        out_dir:str
    ):
        Logger.print("Start create temporary xlsx file with data")
        wb = load_workbook(template_path)
        sheet = wb.active
        Templator.replace_variables_in_sheet(sheet, params)

        result_path:str = os.path.join(
            os.path.abspath(out_dir),
            "temp.xlsx"
        )
        try:
            wb.save(result_path)
        except Exception as e:
            Logger.print(f"Error saving temporary file in {result_path}", level = 'criticla')
            wb.close
            raise FileExistsError(f"Error saving temporary file in {result_path}")
        
        wb.close
        return result_path
    
    @staticmethod
    def replace_variables_in_sheet(
        sheet,
        params: Dict[str, Any]
    ):
        Logger.print("Replacing variables in sheet")
        dict_queue = []

        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for key, value in params.items():
                        placeholder = f'{{{{{key}}}}}'
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(
                                placeholder,
                                str(value)
                            )
                        elif isinstance(value, dict):
                            dict.queue.append(
                                (cell.row, 
                                 cell.column,
                                 value
                                 )
                            )
                            cell.value = cell.value.replace(placeholder, '')
                            break
                        else:
                            Logger.print(f"Error replace variables sheet with key {key}{value}")
                            raise ValueError(f"Error replace variables sheet with key {key}{value}")
        
        for row, col, value in dict_queue:
            Templator.insert_table(sheet, row, col, value)
        
        Logger.print("Finish replacing variables in sheet")

    @staticmethod
    def insert_table(
        sheet,
        start_row:int,
        start_col:int,
        table_data: Dict[str, Any]
    ) -> None:
        Logger.print(f"Replacing tables in {sheet}:{start_row},{start_col}")
        columns = list(table_data.values())
        num_rows = max(len(col) for col in columns)

        # Hash error when try copy style like
        # target.style = copy(source.style)
        def copy_style(
                source: Cell,
                target: Cell
        ):
            try:
                if source.has_style:
                    target.fill = PatternFill(
                        fill_type=source.fill.fill_type,
                        start_color=source.fill.start_color,
                        end_color=source.fill.end_color
                    )
                    target.font = Font(
                        name=source.font.name,
                        size=source.font.size,
                        bold=source.font.bold,
                        italic=source.font.italic,
                        underline=source.font.underline,
                        strike=source.font.strike,
                        color=source.font.color
                    )
                    target.alignment = Alignment(
                        horizontal=source.alignment.horizontal,
                        vertical=source.alignment.vertical,
                        wrap_text=source.alignment.wrap_text,
                        shrink_to_fit=source.alignment.shrink_to_fit,
                        indent=source.alignment.indent,
                    )
                    target.border = Border(
                        left=source.border.left,
                        right=source.border.right,
                        top=source.border.top,
                        bottom=source.border.bottom,
                        diagonal=source.border.diagonal,
                        diagonal_direction=source.border.diagonal_direction,
                        diagonalDown=source.border.diagonalDown,
                        diagonalUp=source.border.diagonalUp,
                        vertical=source.border.vertical,
                        horizontal=source.border.horizontal,
                        outline=source.border.outline,
                        start=source.border.start,
                        end=source.border.end,
                    )
                    target.side = Side (
                        style=source.side.style,
                        border_style=source.side.border_style
                    )

                    target.protection = Protection(
                        locked=source.protection.locked,
                        hidden=source.protection.hidden,
                    )

            except Exception as e:
                Logger.print(f"Error {e}", level="warning")

        original_merges = []

        for i, x in enumerate(sheet.merged_cells.ranges):
            original_merges.append(str(x))

        for i in original_merges:
            sheet.unmerge_cells(i)

        for row_idx in range(num_rows):
            if row_idx == 0:
                sheet.insert_row(start_row + start_col)
            
            for col_idx, col_data in enumerate(columns):
                cell = sheet.cell(
                    row=start_row,
                    columns=start_col + col_idx
                )
                cell.value = col_data[row_idx] if row_idx < len(col_data) else ''
                copy_style(
                    sheet.cell(row=start_row - 1, column = start_col + col_idx),
                    cell
                )

            sheet.row_dimensions[row_idx].height = sheet.row_dimensions[row_idx-1].height

        for org_merg in original_merges:
            numeric = range_boundaries(str(org_merg))
            if(numeric[1] > start_row):
                start_row_new = numeric[1] + num_rows - 1
                start_column_new = num_rows[0]
                end_row_new = numeric[3] + num_rows - 1
                end_column_new = numeric[2]
            else:
                start_row_new = numeric[1]
                start_column_new = numeric[0]
                end_row_new = numeric[3]
                end_column_new = numeric[2]
            sheet.merge_cells(
                start_row=start_row_new,
                start_column=start_column_new,
                end_row=end_row_new,
                emd_column=end_column_new
            )
        Logger.print("End replacing tables")

    @staticmethod
    def excell_to_pdf(
        xlsx_file:str,
        out_dir:str,
        libreoffice_path:str
    ) -> str:
        excell_file = os.path.abspath(xlsx_file)
        command = [
            libreoffice_path,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir',
            os.path.abspath(out_dir),
            excell_file
        ]

        try:
            subprocess.run(
                command,
                capture_output=True,
                shell=True,
                check=True
            )
        except subprocess.CalledProcessError as e:
            Logger.print(f"Error converting xlsx to pdf: {e}", level='error')
            raise FileExistsError(f"Error converting xlsx to pdf {e}")
        
        return str(excell_file).replace("xlsx", "pdf")
    
    @staticmethod
    def add_metadata_to_pdf(
        input_pdf: str,
        output_pdf: str,
        watermark_pdf: str,
        metadata: Dict[str, str]
    ) -> None:
        reader = PdfReader(input_pdf)
        writer = PdfWriter()
        writer.add_metadata(metadata)

        if(watermark_pdf):
            watermark_reader = PdfReader(watermark_pdf)
            watermark_pdf: PageObject = watermark_reader.pages[0]

            for page in reader.pages:
                page.merge_page(watermark_pdf)
                writer.add_page(page)
        with open(output_pdf, 'wb') as f:
            writer.write(f)